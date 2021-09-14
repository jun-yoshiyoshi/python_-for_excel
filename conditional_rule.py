#セルの色付け

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

wb = load_workbook("売上実績.xlsx")
ws = wb.active

red_fill = PatternFill(bgColor='FF0000', fill_type='solid')
cell_rule = CellIsRule(operator='greaterThanOrEqual',
                       formula=[300],
                       fill=red_fill)
#CellIsRuleオブジェクトで色を付ける条件を設定
ws.conditional_formatting.add('F3:F12', cell_rule)
#CellIsRuleはシートに対して設定するので、範囲指定は不要。

wb.save("売上実績_変更後.xlsx")
