# テンプレート入力欄を初期化する
from datetime import date
from openpyxl import load_workbook

wb = load_workbook("申請書xlsx")
ws = wb.active

ws["C4"].value = "営業一課"
ws["C5"].value = "佐藤"
ws["C6"].value = date.today()

for row in ws.iter_rows(min_row=11, max_row=ws.max_row, min_col=2, max_col=7):
    row[0].value = None
    row[4].value = None
    row[5].value = None

wb.save("申請書_返還後.xlsx")
