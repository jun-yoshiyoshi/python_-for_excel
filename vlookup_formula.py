#VLOOKUPに依存するデータの読み込み

from openpyxl import load_workbook

wb = load_workbook("作業時間表.xlsx")

lastmonth = "202004"
month = "202005"
#"202004"と"202005"は同一ブック内の異なるシート

ws_lastmonth = wb[lastmonth]
ws = wb[month]

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    row_count = row[0].row
    row[4].value = f'=VLOOKUP({row_count},{lastmonth}!$B$2:$D$11,3,FALSE)'
    #Ｂ列に設定されている名前が前月シートのセルＢ２～Ｄ１１に該当するとき、Ｄ列の値を取得するＶＬＯＯＫＵＰ関数。

wb.save("作業時間表_変更後.xlsx")