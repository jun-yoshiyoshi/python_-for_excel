#相対参照で数式を縦方向へ貼り付ける

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

wb = load_workbook('粗利.xlsx')
ws = wb.active

#D列が空白。セルCからセルBの値を引く数式をD列全てに入力したい。
origin_cell_no = 'D6'
ws[origin_cell_no] = 'C6-B6'

for row_no in range(7, ws.max_row + 1):
    #６行目と同様の処理を７行目から最終行まで繰り返す。
    cell_no = f"D{row_no}"

    ws[cell_no] = Translator(ws[origin_cell_no].value,
                             origin=origin_cell_no).translate_formula(cell_no)
    #セル情報を数式のままD列へ入力する。

wb.save('粗利_変更後.xlsx')