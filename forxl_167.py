# 異なるシートから情報を取得して計算する

from openpyxl import load_workbook

wb = load_workbook("売上実績.xlsx")
ws_syukei = wb["集計"]

row_list = []
# 集計シートが先頭にあるので、２番目以降のシートのデータを取得する
for ws in wb.worksheets[1:]:
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=3, max_col=ws.max_column):
        cell_list = [(cell.value, cell.number_format) for cell in row]
        # 行に含まれるセルの値とnumber_format属性（表示形式）をタプル型のセットにしてリスト化する。
        row_list.append(cell_list)


# データシートに順番に書き込む処理。表示形式は引き継がれる。
for i, row in enumerate(row_list):
    row_no = i + 3
    ws_syukei.cell(row_no, 2).value = i + 1
    for j, cell in enumerate(row):
        col_no = j + 3
        ws_syukei.cell(row_no, col_no).value = cell[0]
        ws_syukei.cell(row_no, col_no).number_format = cell[1]

wb.save("売上実績_集計後.xlsx")
