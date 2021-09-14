#コマンドライン引数を使ってパラメーター表示する方法

import sys

from openpyxl import load_workbook

filename = sys.argv[1]

cellno = sys.argb[2]

#第１引数のファイル名と第２引数のセル名で出力値を特定。

wb = load_workbook(filename, read_only=True)
ws = wb.active

print(ws[cellno].value)