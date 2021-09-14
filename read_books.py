#複数のエクセルブックから必要な情報を読み取る

from openpyxl import Workbook, load_workbook
from pathlib import Path

count = input("作成するシート数:")

wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = '一覧表'

ws_new["B2"] = "部署名"
ws_new["C2"] = "氏名"

path = Path("./books")

for i, file in enumerate(path.glob("*.xlsx")):
    wb = load_workbook(file, read_only=True)  #reado_only=Trueで元データへの変更を禁止
    ws = wb["チェックリスト"]
    row_no = i + 3
    ws_new[f"B{row_no}"] = ws["C2"].value
    ws_new[f"C{row_no}"] = ws["C3"].value

wb.save("一覧表.xlsx")