from openpyxl import load_workbook
import os
import time
import win32api
import win32print

wb = load_workbook("テスト.xlsx")
# すべてのシート名を取得
wsall = wb.sheetnames
# すべてのシートをtablesectedをFalseにしてアクティブなシートは０番目であると明示する。そうしなければ、意図しないシートが印刷される
# for文ですべてのシートtablesectedプロパティをFalseにしてどのシートも選択してない状態をつくる
for s in wsall:
    ws = wb[s]
    ws.sheet_view.tabSelected = False
wb.save("printed_file.xlsx")
wb.close()
# どのシートも選択してない状態が完成。そして再び開く
wb = load_workbook("printed_file.xlsx")
# これで、0番目＝一番左のシートが選択された状態で開く。for文やif文で印刷したい範囲を指定する。
# 例えば、3番目と4番目シートを印刷する場合range(2,4)
for i in range(0):
    ws = wb.worksheets[i]
    wb.active = i
    ws.sheet_view.tabSelected = True
wb.save("printed_file.xlsx")
# 選択した状態で一度保存。保存しないと保存前の選択状態で印刷されてしまう。
# 印刷する関数 win32api.ShellExecuteを使用してデフォルトプリンターから出力する

folder = os.getcwd()
filename = os.path.join(folder, "printed_file.xlsx")


def PrintOut():
    win32api.ShellExecute(
        0,
        "print",
        filename,
        "/c:""%s" % win32print.GetDefaultPrinter(),
        ".",
        0
    )


PrintOut()
# ワークブックを閉じる
wb.close()
print('実行されました')
time.sleep(1)
