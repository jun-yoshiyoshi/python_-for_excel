#条件を満たす行の色を変更する
import openpyxl#openpyxlをインポート
from openpyxl.styles import PatternFill,Border,Side,Alignment,Protection,Font
#売上金額が35万円以上のセルのフォントを定義。#赤、緑、青の各色を16進法で00からFFまでの各調で表現します。細かく色を指定するには「16進法 カラーコード表」を検索。
high_font=Font(name='MS Pゴシック',size=14,bold=True,italic=True,vertAlign=None,underline='none',strike=False,color='3cb371')
#数値の書式を定義
number_format='\#,##0;\-#,##0'
#ファイルを開く
workBook=openpyxl.load_workbook("新宿店売上リスト.xlsx")#エクセルブックを開く
sheet=workBook["新宿店"]#「新宿店」シートを選択
for row in sheet:#行ループ
	for cell in row:#列ループ
		if sheet.cell(row=1,column=cell.column).value=='売上金額':#もしも列が売上金額のとき
			sales_column=cell.column#売上金額のセル位置を格納
			cell.number_format=number_format#セルに数値形式の書式を設定			
		if cell.row!=1 and sheet.cell(row=cell.row,column=sales_column).value>=350000:#セルの値が35万円以上のとき		
			cell.font=high_font#セルのフォントを設定
workBook.save('新宿店売上リスト(完成)2.xlsx')#ファイル名を指定して保存