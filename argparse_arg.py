#argparseモジュール
# 参照１　https://qiita.com/kzkadc/items/e4fc7bc9c003de1eb6d0
# 参照２　https://docs.python.org/ja/3/library/argparse.html

import argparse
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description='Excelのセルの値を取得するアプリ')

parser.add_argument('filename', help='読み込むブック名:')
parser.add_argument('cellno', help='読み込むセル名（例 A1）:')

args = parser.parse_args()

wb = load_workbook(args.filename, read_only=True)
ws = wb.active

print(ws[args.cellno].value)