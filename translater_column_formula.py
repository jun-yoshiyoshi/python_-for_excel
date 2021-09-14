#数式を相対参照で横方向に貼り付ける

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

wb = load_workbook('合計.xlsx')
ws = wb.active

#セルC3に列CのSUM関数が設定されていて、セルD3が空白。
# セルD3に列DのSUN関数の数式を入力。
ws['D3'] = Translator(ws['C3'].value, origin='C3').translate_formula('D3')

wb.save('合計_変更後.xlsx')