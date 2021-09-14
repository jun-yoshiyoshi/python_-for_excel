#特定の名前のシートを移動する（シートを末尾へ移動）

from openpyxl import load_workbook

wb = load_workbook("チェックリスト.xlsx")

wb.move_sheet('まとめ', offset=len(wb.sheetnames))
#"まとめ"シートを末尾へ移動

wb.save("チェックリスト_変更後.xlsx")
