#　configparser(パラメーター情報を設定ファイル"%.ini"から読み込む方法)
#参照１　https://qiita.com/mimitaro/items/3506a444f325c6f980b2
#参照２　https://docs.python.org/ja/3/library/configparser.html

import configparser
from openpyxl import load_workbook

config = configparser.ConfigParser()

config.read('sample.ini', encoding='utf-8')
#設定ファイル"sample.ini"からパラメーターを読み込む

default = config['DEFAULT']
filename = default['filename']
cellno = default['cellno']

wb = load_workbook(filename, read_only=True)
ws = wb.active

print(ws[cellno].value)
