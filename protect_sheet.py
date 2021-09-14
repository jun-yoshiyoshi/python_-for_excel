#セルを保護する

from openpyxl import load_workbook
from openpyxl.style import Protection

wb = load_workbook("見積書.xlsx")
ws = wb["見積書"]

for rows in ws["B11:H24"]:
    for cell in rows:
        cell.protection = Protection(locked=False)

wb.protecetion.password = "test"
ws.protection.enable()

wb.save("見積書_範囲保護.xlsx")