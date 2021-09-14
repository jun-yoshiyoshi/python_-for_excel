#条件を満たす行に色を付ける
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

wb = load_workbook("課題一覧.xlsx")
ws = wb.active

gray_fill = PatternFill(bgColor='C0C0C0', fill_type="solid")
#色設定の変数をつくる。
cell_rule = FormulaRule(formula=['$D3="完了"'], fill=gray_fill)
#条件付き書式の変数をつくる。"$D"でD列を固定。D4,D5,...が対象となる。
ws.conditional_formatting.add("B3:G12", cell_rule)
#第１引数でセル範囲、第２引数で条件付き書式。
wb.save("課題一覧_変更後.xlsx")