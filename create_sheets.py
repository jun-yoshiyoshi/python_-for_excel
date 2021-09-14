#連番の複数のシートを作成する

from openpyxl import Workbook

count = input("作成するシート数:")

wb = Workbook()
ws = wb.active
ws.title = '概要_1'

for i in range(2, int(count) + 1):
    wb.create_sheet(title=f"概要_{i}")
    #第二引数で新規シートの位置を指定できる。
    #wb.create_sheet(title=f"概要_{i}",index=0)ブックの先頭に作成

wb.save("資料.xlsx")